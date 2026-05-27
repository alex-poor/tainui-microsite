#!/usr/bin/env python3
"""Extract MSD hardship data from Excel to clean CSVs for the microsite."""

import openpyxl
import csv
import os

SRC = os.path.join(os.path.dirname(__file__), '..', 'assets', 'BIIM-7241_Waikato_hardship_.xlsx')
OUT = os.path.join(os.path.dirname(__file__), '..', 'data')
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)

QUARTERS = [
    'Mar 2023', 'Jun 2023', 'Sep 2023', 'Dec 2023',
    'Mar 2024', 'Jun 2024', 'Sep 2024', 'Dec 2024',
    'Mar 2025', 'Jun 2025', 'Sep 2025', 'Dec 2025',
    'Mar 2026'
]

HARDSHIP_TYPES = {
    'Electricity Assistance',
    'Emergency Housing',
    'Food',
    'Gas Assistance',
    'Stranded Travel - Petrol',
    'All hardship',
}


def extract_tla_sheet(sheet_name, value_col, out_file):
    """TLA sheets: col0=TLA or hardship type, col1=hardship type, cols 2-14=quarters."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Find the header row with quarter names
    header_row_idx = None
    for i, row in enumerate(rows):
        if row[2] and 'March 2023' in str(row[2]):
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"  ERROR: could not find header row in {sheet_name}")
        return

    records = []
    current_tla = None

    for row in rows[header_row_idx + 1:]:
        cell0 = str(row[0]).strip() if row[0] else ''
        cell1 = str(row[1]).strip() if row[1] else ''

        if not cell0 and not cell1:
            continue
        if cell0.startswith('Source') or cell0.startswith('Note'):
            break

        # TLA name in col0, hardship type in col1
        if cell0 and cell0 not in HARDSHIP_TYPES:
            current_tla = cell0

        hardship = cell1 if cell1 in HARDSHIP_TYPES else (cell0 if cell0 in HARDSHIP_TYPES else None)
        if not hardship or not current_tla:
            continue

        for qi, q in enumerate(QUARTERS):
            val = row[qi + 2]
            if val is not None and str(val).strip() != '':
                records.append({
                    'tla': current_tla,
                    'hardship_type': hardship,
                    'quarter': q,
                    value_col: val
                })

    with open(os.path.join(OUT, out_file), 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['tla', 'hardship_type', 'quarter', value_col])
        writer.writeheader()
        writer.writerows(records)

    print(f"  {out_file}: {len(records)} rows")


def extract_region_sheet(sheet_name, value_col, out_file):
    """Region sheets: col0=region, col1=service centre, col2=hardship type, cols 3-15=quarters."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    for i, row in enumerate(rows):
        if row[3] and 'March 2023' in str(row[3]):
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f"  ERROR: could not find header row in {sheet_name}")
        return

    records = []
    current_region = None
    current_centre = None

    for row in rows[header_row_idx + 1:]:
        cell0 = str(row[0]).strip() if row[0] else ''
        cell1 = str(row[1]).strip() if row[1] else ''
        cell2 = str(row[2]).strip() if row[2] else ''

        if not cell0 and not cell1 and not cell2:
            continue
        if cell0.startswith('Source') or cell0.startswith('Note'):
            break

        if cell0 and cell0 not in HARDSHIP_TYPES and cell0 != 'None':
            current_region = cell0
        if cell1 and cell1 not in HARDSHIP_TYPES and cell1 != 'None':
            current_centre = cell1

        hardship = cell2 if cell2 in HARDSHIP_TYPES else None
        if not hardship or not current_region or not current_centre:
            continue

        for qi, q in enumerate(QUARTERS):
            val = row[qi + 3]
            if val is not None and str(val).strip() != '' and str(val) != 'None':
                records.append({
                    'region': current_region,
                    'service_centre': current_centre,
                    'hardship_type': hardship,
                    'quarter': q,
                    value_col: val
                })

    with open(os.path.join(OUT, out_file), 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['region', 'service_centre', 'hardship_type', 'quarter', value_col])
        writer.writeheader()
        writer.writerows(records)

    print(f"  {out_file}: {len(records)} rows")


print("Extracting TLA sheets...")
extract_tla_sheet('TLA_grant', 'grants', '01_tla_grants.csv')
extract_tla_sheet('TLA_amount', 'amount', '02_tla_amounts.csv')

print("Extracting Region sheets...")
extract_region_sheet('Region_grant', 'grants', '03_region_grants.csv')
extract_region_sheet('Region_amount', 'amount', '04_region_amounts.csv')

print("\nDone! CSVs written to data/")
